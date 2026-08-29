"""Multi-tab .xlsx workbook — the portable form of the 6-tab deliverable.

Google's "secure by default" org policy blocks service-account key creation on
new accounts, so the direct Sheets API push (`SheetsSink`) often isn't
available. This sink writes every tab into one workbook at
`data/output/<name>.xlsx`; the reviewer uploads it to Google Drive and opens it
as a Google Sheet (or just opens the file), giving the same 6 tabs:

    Startups | Products | Research Papers | Jobs | News | Entity Mapping Log

Same `write_records` / `write_rows` interface as `SheetsSink` so the pipeline
writes to both without caring which one is active. Call `save()` once at the end.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from src.config import get_pipeline_config
from src.logging_setup import get_logger
from src.sinks.sheets_sink import _flatten  # reuse the dotted-key flattener

log = get_logger("sink")


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int | float | str):
        return v
    return str(v)


class XlsxSink:
    def __init__(self, path: str | None = None) -> None:
        out = get_pipeline_config()["output"]
        self._tab_names: dict[str, str] = out["sheets"]["tabs"]
        self._path = Path(path or Path(out["jsonl_dir"]) / "graphone_intelligence.xlsx")
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._wb = Workbook()
        self._wb.remove(self._wb.active)  # drop the default empty sheet
        self._written: dict[str, int] = {}

    # ------------------------------------------------------------------ #
    def write_records(self, tab_key: str, records: Iterable[BaseModel], columns: list[str]) -> None:
        rows = [[_flatten(r.model_dump(mode="json")).get(c) for c in columns] for r in records]
        self._sheet(tab_key, columns, rows)

    def write_rows(self, tab_key: str, rows: list[dict], columns: list[str]) -> None:
        matrix = [[r.get(c) for c in columns] for r in rows]
        self._sheet(tab_key, columns, matrix)

    # ------------------------------------------------------------------ #
    def _sheet(self, tab_key: str, columns: list[str], rows: list[list[Any]]) -> None:
        title = self._tab_names.get(tab_key, tab_key)[:31]  # Excel tab name limit
        ws = self._wb[title] if title in self._wb.sheetnames else self._wb.create_sheet(title)
        ws.delete_rows(1, ws.max_row)
        ws.append(columns)
        for row in rows:
            ws.append([_cell(v) for v in row])
        for i, col in enumerate(columns, start=1):
            width = max(len(col), *(len(str(_cell(r[i - 1]))) for r in rows)) if rows else len(col)
            ws.column_dimensions[get_column_letter(i)].width = min(60, width + 2)
        ws.freeze_panes = "A2"
        self._written[title] = len(rows)

    def save(self) -> Path:
        # stable tab order matching the brief
        order = ["startups", "products", "research_papers", "jobs", "news", "mapping_log"]
        desired = [self._tab_names[k] for k in order if k in self._tab_names]
        self._wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)
        self._wb.save(self._path)
        log.info("xlsx_saved", path=str(self._path), tabs=self._written)
        return self._path
